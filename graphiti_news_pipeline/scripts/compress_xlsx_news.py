#!/usr/bin/env python3
"""
Compress news content in an XLSX file with LLM and write a new compressed XLSX.

Default behavior:
- read first non-compressed .xlsx from ./docs
- compress `content` to <= 100 characters
- write `<input>.compressed.xlsx`
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from dotenv import load_dotenv
from openpyxl import load_workbook


@dataclass
class LLMSettings:
    api_key: str
    api_base: str
    model: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compress XLSX news content with LLM and output a compressed XLSX file."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=None,
        help="Input xlsx path. Defaults to first docs/*.xlsx excluding *.compressed.xlsx.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output xlsx path. Defaults to <input>.compressed.xlsx.",
    )
    parser.add_argument(
        "--max-chars",
        type=int,
        default=100,
        help="Maximum characters for compressed content (default: 100).",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Maximum rows to compress (0 means all rows).",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=600,
        help="Timeout seconds for each LLM request (default: 600).",
    )
    parser.add_argument(
        "--retries",
        type=int,
        default=2,
        help="Retry times for each row when compression fails (default: 2).",
    )
    parser.add_argument(
        "--interval-seconds",
        type=float,
        default=0.5,
        help="Sleep interval between row compress requests (default: 0.5).",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite output if exists.",
    )
    parser.add_argument(
        "--continue-on-error",
        action="store_true",
        help="Keep original content for failed rows and continue.",
    )
    return parser.parse_args()


def pick_default_input_xlsx() -> Path:
    files = sorted(Path("docs").glob("*.xlsx"))
    filtered = [p for p in files if not p.name.endswith(".compressed.xlsx")]
    if not filtered:
        raise FileNotFoundError("No input xlsx found under ./docs (excluding *.compressed.xlsx).")
    return filtered[0]


def default_output_path(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}.compressed{input_path.suffix}")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def format_publish_date(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def post_json(
    url: str,
    payload: dict[str, Any],
    timeout: int,
    headers: dict[str, str] | None = None,
) -> tuple[int, str]:
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req_headers = {"Content-Type": "application/json; charset=utf-8"}
    if headers:
        req_headers.update(headers)
    req = Request(url=url, data=body, headers=req_headers, method="POST")
    try:
        with urlopen(req, timeout=timeout) as resp:
            raw = resp.read().decode("utf-8", errors="replace")
            return resp.status, raw
    except HTTPError as e:
        raw = e.read().decode("utf-8", errors="replace")
        return e.code, raw
    except URLError as e:
        raise RuntimeError(f"Network error: {e}") from e


def get_llm_settings() -> LLMSettings:
    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    api_base = os.getenv("OPENAI_API_BASE", "").strip()
    model = os.getenv("OPENAI_MODEL", "").strip()
    if not api_key or not api_base or not model:
        raise RuntimeError("Missing OPENAI_API_KEY/OPENAI_API_BASE/OPENAI_MODEL in environment.")
    return LLMSettings(api_key=api_key, api_base=api_base.rstrip("/"), model=model)


def cleanup_llm_text(text: str, max_chars: int) -> str:
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = cleaned.strip("`").strip()
    if len(cleaned) > max_chars:
        cleaned = cleaned[:max_chars].rstrip()
    return cleaned


def compress_with_llm(
    llm: LLMSettings,
    title: str,
    publish_time: str,
    content: str,
    timeout: int,
    max_chars: int,
) -> str:
    endpoint = f"{llm.api_base}/chat/completions"
    payload = {
        "model": llm.model,
        "temperature": 0,
        "max_tokens": 512,
        "messages": [
            {
                "role": "system",
                "content": (
                    "You compress Chinese news text for downstream ingestion. "
                    "Keep only core facts: key entities, main event, time, place, and key numbers. "
                    "Do not hallucinate. Output plain Chinese text only."
                ),
            },
            {
                "role": "user",
                "content": (
                    f"Compress the following news to <= {max_chars} Chinese characters.\n"
                    f"Title: {title}\n"
                    f"Publish time: {publish_time or 'unknown'}\n"
                    f"Body:\n{content}"
                ),
            },
        ],
    }
    headers = {"Authorization": f"Bearer {llm.api_key}"}
    status, raw = post_json(endpoint, payload, timeout=timeout, headers=headers)
    if not (200 <= status < 300):
        raise RuntimeError(f"LLM compression failed status={status} body={raw[:500]}")
    try:
        data = json.loads(raw)
        message = data["choices"][0]["message"]["content"]
    except Exception as exc:
        raise RuntimeError(f"Invalid LLM response format: {raw[:500]}") from exc
    result = cleanup_llm_text(str(message), max_chars=max_chars)
    if not result:
        raise RuntimeError("LLM returned empty compressed content.")
    return result


def main() -> int:
    args = parse_args()
    sys.stdout.reconfigure(encoding="utf-8")
    load_dotenv()

    llm = get_llm_settings()
    input_path = args.input if args.input is not None else pick_default_input_xlsx()
    output_path = args.output if args.output is not None else default_output_path(input_path)

    if not input_path.exists():
        print(f"[ERROR] input xlsx not found: {input_path}")
        return 1
    if output_path.exists() and not args.overwrite:
        print(f"[ERROR] output exists: {output_path}. Use --overwrite to replace.")
        return 1

    wb = load_workbook(input_path)
    ws = wb.active

    header_row = [cell.value for cell in ws[1]]
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    index_map = {name: idx for idx, name in enumerate(headers)}

    if "content" not in index_map:
        print("[ERROR] missing required column: content")
        return 1

    title_idx = index_map.get("title")
    publish_idx = index_map.get("publish_date")
    content_idx = index_map["content"]
    raw_idx = index_map.get("raw_content")

    if raw_idx is None:
        raw_idx = ws.max_column
        ws.cell(row=1, column=raw_idx + 1, value="raw_content")
        raw_idx = raw_idx

    max_rows = ws.max_row
    processed = 0
    success = 0
    failed = 0
    skipped_short = 0

    print(f"[INFO] input={input_path}")
    print(f"[INFO] output={output_path}")
    print(f"[INFO] max_rows={max_rows - 1}")
    print(f"[INFO] max_chars={args.max_chars}")

    for row_no in range(2, max_rows + 1):
        if args.limit > 0 and processed >= args.limit:
            break

        content_cell = ws.cell(row=row_no, column=content_idx + 1)
        content = normalize_text(content_cell.value)
        if not content:
            continue

        processed += 1

        title = ""
        if title_idx is not None:
            title = normalize_text(ws.cell(row=row_no, column=title_idx + 1).value)
        publish_time = ""
        if publish_idx is not None:
            publish_time = format_publish_date(ws.cell(row=row_no, column=publish_idx + 1).value)

        raw_cell = ws.cell(row=row_no, column=raw_idx + 1)
        if not normalize_text(raw_cell.value):
            raw_cell.value = content

        if len(content) <= args.max_chars:
            skipped_short += 1
            print(f"[{processed}] row={row_no} skip len={len(content)}")
            continue

        print(f"[{processed}] row={row_no} compress len={len(content)}", end="")
        last_error: Exception | None = None
        compressed: str | None = None

        for attempt in range(args.retries + 1):
            try:
                compressed = compress_with_llm(
                    llm=llm,
                    title=title,
                    publish_time=publish_time,
                    content=content,
                    timeout=args.timeout,
                    max_chars=args.max_chars,
                )
                break
            except Exception as exc:  # noqa: BLE001
                last_error = exc
                if attempt < args.retries:
                    time.sleep(1.0)

        if compressed is None:
            failed += 1
            print(" -> FAIL")
            print(f"    error: {last_error}")
            if not args.continue_on_error:
                print("[ERROR] stop on first compression failure.")
                return 2
            continue

        content_cell.value = compressed
        success += 1
        print(f" -> OK new_len={len(compressed)}")

        if args.interval_seconds > 0:
            time.sleep(args.interval_seconds)

    wb.save(output_path)
    print(
        "[SUMMARY] "
        f"processed={processed} success={success} failed={failed} skipped_short={skipped_short}"
    )
    return 0 if failed == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
