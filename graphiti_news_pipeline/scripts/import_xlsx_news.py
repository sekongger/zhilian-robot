#!/usr/bin/env python3
"""
Import rows from an XLSX file and send them to /api/add-text.

Default behavior:
- find the first .compressed.xlsx file under ./docs
- import first 10 valid rows
- call http://localhost:8000/api/add-text
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from dotenv import load_dotenv
from openpyxl import load_workbook


@dataclass
class NewsRow:
    row_number: int
    title: str
    publish_date: str
    source: str
    url: str
    content: str
    raw_content: str


@dataclass
class LLMSettings:
    api_key: str
    api_base: str
    model: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import news rows from xlsx into graphiti add-text endpoint."
    )
    parser.add_argument(
        "--file",
        type=Path,
        default=None,
        help="Path to xlsx file. Defaults to first docs/*.compressed.xlsx.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=10,
        help="Maximum number of rows to import (default: 10).",
    )
    parser.add_argument(
        "--api-base",
        type=str,
        default="http://localhost:8000/api",
        help="API base url (default: http://localhost:8000/api).",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=600,
        help="Request timeout seconds for both LLM compression and add-text API (default: 600).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Only print rows that would be imported, do not call API.",
    )
    parser.add_argument(
        "--interval-seconds",
        type=float,
        default=10.0,
        help="Sleep interval between each row import request (default: 10).",
    )
    parser.add_argument(
        "--enable-compress",
        action="store_true",
        help="Enable LLM compression before upload (default: disabled).",
    )
    parser.add_argument(
        "--skip-compress",
        action="store_true",
        help="Deprecated compatibility flag. Compression is already disabled by default.",
    )
    return parser.parse_args()


def pick_default_xlsx() -> Path:
    compressed_files = sorted(Path("docs").glob("*.compressed.xlsx"))
    if compressed_files:
        return compressed_files[0]

    files = sorted(Path("docs").glob("*.xlsx"))
    if files:
        raise FileNotFoundError(
            "No compressed xlsx found under ./docs. "
            "Run scripts/compress_xlsx_news.py first or pass --file explicitly."
        )
    raise FileNotFoundError("No xlsx file found under ./docs")


def format_publish_date(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def read_rows(xlsx_path: Path, limit: int) -> list[NewsRow]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    index_map = {name: idx for idx, name in enumerate(headers)}

    required = ["title", "content"]
    missing = [name for name in required if name not in index_map]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    title_idx = index_map["title"]
    content_idx = index_map["content"]
    publish_idx = index_map.get("publish_date")
    source_idx = index_map.get("source")
    url_idx = index_map.get("url")
    raw_content_idx = index_map.get("raw_content")

    rows: list[NewsRow] = []
    for excel_row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        title = normalize_text(row[title_idx] if title_idx < len(row) else "")
        content = normalize_text(row[content_idx] if content_idx < len(row) else "")

        publish_date = ""
        if publish_idx is not None and publish_idx < len(row):
            publish_date = format_publish_date(row[publish_idx])

        source = ""
        if source_idx is not None and source_idx < len(row):
            source = normalize_text(row[source_idx])

        url = ""
        if url_idx is not None and url_idx < len(row):
            url = normalize_text(row[url_idx])

        raw_content = ""
        if raw_content_idx is not None and raw_content_idx < len(row):
            raw_content = normalize_text(row[raw_content_idx])
        if not raw_content:
            raw_content = content

        if not title or not content:
            continue

        rows.append(
            NewsRow(
                row_number=excel_row_num,
                title=title,
                publish_date=publish_date,
                source=source,
                url=url,
                content=content,
                raw_content=raw_content,
            )
        )
        if len(rows) >= limit:
            break

    return rows


def post_json(
    url: str,
    payload: dict[str, Any],
    timeout: int,
    headers: dict[str, str] | None = None,
) -> tuple[int, str]:
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    request_headers = {"Content-Type": "application/json; charset=utf-8"}
    if headers:
        request_headers.update(headers)
    req = Request(
        url=url,
        data=body,
        headers=request_headers,
        method="POST",
    )
    try:
        with urlopen(req, timeout=timeout) as resp:
            raw = resp.read().decode("utf-8", errors="replace")
            return resp.status, raw
    except HTTPError as e:
        raw = e.read().decode("utf-8", errors="replace")
        return e.code, raw
    except URLError as e:
        raise RuntimeError(f"Network error: {e}") from e


def get_llm_settings() -> LLMSettings:
    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    api_base = os.getenv("OPENAI_API_BASE", "").strip()
    model = os.getenv("OPENAI_MODEL", "").strip()
    if not api_key or not api_base or not model:
        raise RuntimeError(
            "Missing OPENAI_API_KEY/OPENAI_API_BASE/OPENAI_MODEL in environment for compression."
        )
    return LLMSettings(api_key=api_key, api_base=api_base.rstrip("/"), model=model)


def compress_with_llm(row: NewsRow, llm: LLMSettings, timeout: int) -> str:
    endpoint = f"{llm.api_base}/chat/completions"
    payload = {
        "model": llm.model,
        "temperature": 0,
        "messages": [
            {
                "role": "system",
                "content": (
                    "You compress Chinese news text for knowledge graph ingestion. "
                    "Keep only core facts: key entities, action/event, time, place, and critical numbers. "
                    "No hallucination. Output concise Chinese plain text only."
                ),
            },
            {
                "role": "user",
                "content": (
                    "Compress the following Chinese news text to <=100 Chinese characters. "
                    "Keep only core facts and avoid commentary.\n"
                    f"Title: {row.title}\n"
                    f"Publish time: {row.publish_date or 'unknown'}\n"
                    f"Body:\n{row.content}"
                ),
            },
        ],
    }
    headers = {"Authorization": f"Bearer {llm.api_key}"}
    status, raw = post_json(endpoint, payload, timeout, headers=headers)
    if not (200 <= status < 300):
        raise RuntimeError(f"LLM compression failed status={status} body={raw[:500]}")

    try:
        data = json.loads(raw)
        message = data["choices"][0]["message"]["content"]
    except Exception as exc:
        raise RuntimeError(f"Invalid LLM response format: {raw[:500]}") from exc

    text = str(message).strip()
    if len(text) > 100:
        text = text[:100].rstrip()
    if not text:
        raise RuntimeError("LLM returned empty compression text.")
    return text


def main() -> int:
    args = parse_args()
    sys.stdout.reconfigure(encoding="utf-8")
    load_dotenv()

    xlsx_path = args.file if args.file is not None else pick_default_xlsx()
    if not xlsx_path.exists():
        print(f"[ERROR] xlsx not found: {xlsx_path}")
        return 1

    rows = read_rows(xlsx_path, args.limit)
    if not rows:
        print("[WARN] no valid rows found to import")
        return 0

    endpoint = args.api_base.rstrip("/") + "/add-text"
    llm_settings = None
    compress_enabled = args.enable_compress and not args.skip_compress

    if compress_enabled and not args.dry_run:
        try:
            llm_settings = get_llm_settings()
        except Exception as exc:
            print(f"[ERROR] failed to load LLM settings: {exc}")
            return 1

    print(f"[INFO] file={xlsx_path}")
    print(f"[INFO] endpoint={endpoint}")
    print(f"[INFO] rows_to_import={len(rows)}")
    print(f"[INFO] compression={'on' if compress_enabled else 'off'}")

    success = 0
    failed = 0
    for i, row in enumerate(rows, start=1):
        print(f"[{i}/{len(rows)}] row={row.row_number} title={row.title}")

        if args.dry_run:
            print("  -> dry-run skip")
            continue

        text_to_send = row.content
        if llm_settings is not None:
            try:
                compressed = compress_with_llm(row, llm_settings, args.timeout)
                text_to_send = compressed
                print(f"  -> compressed chars={len(compressed)}")
            except Exception as exc:
                failed += 1
                print(f"  -> FAIL compress: {exc}")
                continue

        payload = {
            "title": row.title,
            "name": row.title,
            "text": text_to_send,
            "publish_time": row.publish_date or None,
            "source": row.source or None,
            "url": row.url or None,
            "raw_text": row.raw_content,
        }

        try:
            status, raw = post_json(endpoint, payload, args.timeout)
        except Exception as exc:
            failed += 1
            print(f"  -> FAIL exception: {exc}")
            continue

        if 200 <= status < 300:
            success += 1
            print(f"  -> OK status={status}")
        else:
            failed += 1
            print(f"  -> FAIL status={status} body={raw[:500]}")

        if i < len(rows) and args.interval_seconds > 0:
            print(f"  -> sleep {args.interval_seconds}s before next row")
            time.sleep(args.interval_seconds)

    print(f"[SUMMARY] success={success} failed={failed} total={len(rows)}")
    return 0 if failed == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
